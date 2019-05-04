import tkinter as tk
import cv2
from PIL import Image, ImageTk
import face_recognition
import time
from imutils.video import VideoStream
import numpy as np
from copy import deepcopy
from openpyxl import *


class CamView:
    def __init__(self, parent):
        self.parent = parent
        self.window = tk.Toplevel(parent)

        self.lmain2 = tk.Label(self.window)
        self.lmain2.pack()

        self.window.protocol("WM_DELETE_WINDOW", self.close)
        self.show_frame()

    def show_frame(self):
        imgtk = ImageTk.PhotoImage(image=self.parent.img)
        self.lmain2.imgtk = imgtk
        self.lmain2.configure(image=imgtk)

    def close(self):
        self.parent.test_frame = None
        self.window.destroy()


class AfterCamView:
    def __init__(self, parent):
        self.parent = parent
        self.window = tk.Toplevel(parent)

        self.lmain2 = tk.Label(self.window)
        self.lmain2.pack()

        self.window.protocol("WM_DELETE_WINDOW", self.close)
        self.wb = load_workbook('excel.xlsx')
        self.sheet = self.wb.active
        self.show_frame()

    def excel(self):
        # resize the width of columns in
        # excel spreadsheet
        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 10
        self.sheet.column_dimensions['C'].width = 10
        self.sheet.column_dimensions['D'].width = 20
        self.sheet.column_dimensions['E'].width = 20
        self.sheet.column_dimensions['F'].width = 40
        self.sheet.column_dimensions['G'].width = 50

        # write given data to an excel spreadsheet
        # at particular location
        self.sheet.cell(row=1, column=1).value = "Name"
        self.sheet.cell(row=1, column=2).value = "Course"
        self.sheet.cell(row=1, column=3).value = "Semester"
        self.sheet.cell(row=1, column=4).value = "Form Number"
        self.sheet.cell(row=1, column=5).value = "Contact Nmber"
        self.sheet.cell(row=1, column=6).value = "Email id"
        self.sheet.cell(row=1, column=7).value = "Address"

    # Function to set focus (cursor)
    def focus1(self):
        # set focus on the course_field box
        self.course_field.focus_set()

    # Function to set focus
    def focus2(self):
        # set focus on the sem_field box
        self.sem_field.focus_set()

    # Function to set focus
    def focus3(self):
        # set focus on the form_no_field box
        self.form_no_field.focus_set()

    # Function to set focus
    def focus4(self):
        # set focus on the contact_no_field box
        self.contact_no_field.focus_set()

    # Function to set focus
    def focus5(self):
        # set focus on the email_id_field box
        self.email_id_field.focus_set()

    # Function to set focus
    def focus6(self):
        # set focus on the address_field box
        self.address_field.focus_set()

    # Function for clearing the
    # contents of text entry boxes
    def clear(self):
        # clear the content of text entry box
        self.name_field.delete(0)
        self.course_field.delete(0)
        self.sem_field.delete(0)
        self.form_no_field.delete(0)
        self.contact_no_field.delete(0)
        self.email_id_field.delete(0)
        self.address_field.delete(0)

    # Function to take data from GUI
    # window and write to an excel file
    def insert(self):
        # if user not fill any entry
        # then print "empty input"
        if (self.name_field.get() == "" and
                self.course_field.get() == "" and
                self.sem_field.get() == "" and
                self.form_no_field.get() == "" and
                self.contact_no_field.get() == "" and
                self.email_id_field.get() == "" and
                self.address_field.get() == ""):
            print("empty input")

        else:

            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = self.sheet.max_row

            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location
            self.sheet.cell(row=current_row + 1, column=1).value = self.name_field.get()
            self.sheet.cell(row=current_row + 1, column=2).value = self.course_field.get()
            self.sheet.cell(row=current_row + 1, column=3).value = self.sem_field.get()
            self.sheet.cell(row=current_row + 1, column=4).value = self.form_no_field.get()
            self.sheet.cell(row=current_row + 1, column=5).value = self.contact_no_field.get()
            self.sheet.cell(row=current_row + 1, column=6).value = self.email_id_field.get()
            self.sheet.cell(row=current_row + 1, column=7).value = self.address_field.get()

            # save the file
            self.wb.save('excel.xlsx')

            # set focus on the name_field box
            self.name_field.focus_set()

            # call the clear() function
            self.clear()

    def show_frame(self):
        self.window.configure(background='light green')

        self.window.title("registration form")

        self.window.geometry("800x500")

        self.excel()

        # create a Form label 
        heading = tk.Label(self.window, text="Form", bg="light green")
        heading.pack()

        # create a Name label
        name = tk.Label(self.window, text="Name", bg="light green")
        name.pack()
        self.name_field = tk.Entry(self.window)
        self.name_field.pack()

        # create a Course label
        course = tk.Label(self.window, text="Course", bg="light green")
        course.pack()
        self.course_field = tk.Entry(self.window)
        self.course_field.pack()

        # create a Semester label
        sem = tk.Label(self.window, text="Semester", bg="light green")
        sem.pack()
        self.sem_field = tk.Entry(self.window)
        self.sem_field.pack()

        # create a Form No. lable
        form_no = tk.Label(self.window, text="Form No.", bg="light green")
        form_no.pack()
        self.form_no_field = tk.Entry(self.window)
        self.form_no_field.pack()

        # create a Contact No. label
        contact_no = tk.Label(self.window, text="Contact No.", bg="light green")
        contact_no.pack()
        self.contact_no_field = tk.Entry(self.window)
        self.contact_no_field.pack()

        # create a Email id label
        email_id = tk.Label(self.window, text="Email id", bg="light green")
        email_id.pack()
        self.email_id_field = tk.Entry(self.window)
        self.email_id_field.pack()

        # create a address label
        address = tk.Label(self.window, text="Address", bg="light green")
        address.pack()
        self.address_field = tk.Entry(self.window)
        self.address_field.pack()

        # bind method of widget is used for
        # the binding the function with the events

        # whenever the enter key is pressed
        # then call the focus1 function
        self.name_field.bind("<Return>", self.focus1)

        # whenever the enter key is pressed
        # then call the focus2 function
        self.course_field.bind("<Return>", self.focus2)

        # whenever the enter key is pressed
        # then call the focus3 function
        self.sem_field.bind("<Return>", self.focus3)

        # whenever the enter key is pressed
        # then call the focus4 function
        self.form_no_field.bind("<Return>", self.focus4)

        # whenever the enter key is pressed
        # then call the focus5 function
        self.contact_no_field.bind("<Return>", self.focus5)

        # whenever the enter key is pressed
        # then call the focus6 function
        self.email_id_field.bind("<Return>", self.focus6)

        # call excel function
        self.excel()

        # create a Submit Button and place into the root window
        submit = tk.Button(self.window, text="Submit", fg="Black", bg="Red", command=self.insert)
        submit.pack()

    def close(self):
        # self.parent.test_frame = None
        self.window.destroy()


class Main(tk.Frame):
    def __init__(self, parent):

        self.lmain = tk.Label(parent)
        self.lmain.pack()

        self.test_frame = None
        frame = tk.Frame.__init__(self, parent)
        tk.Label(text='hello!').pack()
        b = tk.Button(frame, text='open', command=self.load_window)
        b.pack()

        width, height = 800, 600
        self.cap = cv2.VideoCapture(0)
        self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, width)
        self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, height)

        # Load a sample picture and learn how to recognize it.
        self.nikhil_image = face_recognition.load_image_file("nikhil.jpg")
        self.nikhil_face_encoding = face_recognition.face_encodings(self.nikhil_image)[0]

        # Load a second sample picture and learn how to recognize it.
        self.satyam_image = face_recognition.load_image_file("satyam.jpg")
        self.satyam_face_encoding = face_recognition.face_encodings(self.satyam_image)[0]

        # Load a second sample picture and learn how to recognize it.
        self.ramsharan_image = face_recognition.load_image_file("ramsharan.jpg")
        self.ramsharan_face_encoding = face_recognition.face_encodings(self.ramsharan_image)[0]

        # Load a second sample picture and learn how to recognize it.
        self.ravi_image = face_recognition.load_image_file("ravi.jpg")
        self.ravi_face_encoding = face_recognition.face_encodings(self.ravi_image)[0]

        # Create arrays of known face encodings and their names
        self.known_face_encodings = [
            self.nikhil_face_encoding,
            self.satyam_face_encoding,
            self.ramsharan_face_encoding,
            self.ravi_face_encoding
        ]
        self.known_face_names = [
            "Nikhil Kumar",
            "Satyam Singh",
            "Ramsharan Singh",
            "Ravi"
        ]

        # Initialize some variables
        self.face_locations = []
        self.face_encodings = []
        self.face_names = []
        self.process_this_frame = True
        self.last_available_on = time.time()
        self.detectionList = [None] * 10
        self.loginFlag = False

    def do_stuff(self):
        _, frame = self.cap.read()

        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        self.rgb_small_frame = small_frame[:, :, ::-1]

        # Only process every other frame of video to save time
        if self.process_this_frame:
            # print("Inside Processing Frames...")
            # Find all the faces and face encodings in the current frame of video
            self.face_locations = face_recognition.face_locations(self.rgb_small_frame)
            self.face_encodings = face_recognition.face_encodings(self.rgb_small_frame, self.face_locations)

            self.face_names = []
            for face_encoding in self.face_encodings:
                # See if the face is a match for the known face(s)
                matches = face_recognition.compare_faces(self.known_face_encodings, face_encoding)
                name = "Unknown"

                # If a match was found in known_face_encodings, just use the first one.
                print("Matches", matches)

                if True in matches:
                    first_match_index = matches.index(True)
                    name = self.known_face_names[first_match_index]
                    last_available_on = time.time()
                    # print(name, last_available_on);
                else:
                    name = "Unknown"

                self.detectionList = self.detectionList[1:]
                self.detectionList.append(name)

                # print(self.detectionList)
                if self.detectionList.count(name) == 10:
                    self.face_names.append(name)  # Success
                    print("Logged in as", name)
                    self.loginFlag = True
                    if type(self.test_frame) == CamView:
                        self.test_frame.close()
                    elif type(self.test_frame) == AfterCamView:
                        pass
                    else:
                        self.test_frame = AfterCamView(self)
                    break
                else:
                    self.face_names.append("Unknown")


        self.process_this_frame = not self.process_this_frame
        frame2 = deepcopy(frame)

        # Display the results
        for (top, right, bottom, left), name in zip(self.face_locations, self.face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            # Draw a box around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            # Draw a label with a name below the face
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

        # Display the resulting image
        # cv2.imshow('Video', frame)
        # print(np.array_equal(frame,frame2))

        # Hit 'q' on the keyboard to quit!
        # if (cv2.waitKey(1) & 0xFF == ord('q')) or (time.time()-last_available_on>=10):
        if cv2.waitKey(1) & 0xFF == ord('q'):
            return

        # Release handle to the webcam
        # video_capture.release()
        # cv2.destroyAllWindows()

        frame = cv2.flip(frame, 1)
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        self.img = Image.fromarray(cv2image)
        if self.test_frame is not None and type(self.test_frame) != AfterCamView:
            self.test_frame.show_frame()
        self.lmain.after(10, self.do_stuff)

    def load_window(self):
        self.do_stuff()
        if self.test_frame is None:
            self.test_frame = CamView(self)




root = tk.Tk()
root.bind('<Escape>', lambda e: root.quit())
control = Main(root)
root.mainloop()
